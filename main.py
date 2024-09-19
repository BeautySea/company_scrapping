from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, StaleElementReferenceException
import time
import json
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment
import re
import requests

# Path to your ChromeDriver
chrome_driver_path = "./drivers/chromedriver/chromedriver.exe"  # Update this path accordingly
requests.get("https://poit.bolagsverket.se/poit/rest/SokKungorelse?sokord=&kungorelseid=&kungorelseObjektPersonOrgnummer=&kungorelseObjektNamn=&tidsperiod=ANNAN_PERIOD&tidsperiodFrom=2024-09-17&tidsperiodTom=2024-09-17&amnesomradeId=2&kungorelsetypId=4&underRubrikId=6")
# Optional: specify the path to your Chrome executable if it's not in the default location
chrome_options = Options()
chrome_options.binary_location = "./chrome-win64/chrome.exe"
chrome_options.add_argument("--start-maximized")
# chrome_options.add_argument("--headless")  # Ensure headless mode is enabled
chrome_options.add_argument("--no-sandbox")  # Bypass OS security model (necessary for some environments)
chrome_options.add_argument("--disable-dev-shm-usage")  # Overcome limited resource problems
chrome_options.add_argument("--disable-gpu")  # Disable GPU usage for better performance in headless mode
chrome_options.add_argument("--remote-debugging-port=9222")  # Fix for DevToolsActivePort file doesn't exist
chrome_options.add_argument("--disable-software-rasterizer")  # Disables use of software rasterizer


# Set up the ChromeDriver service
service = Service(chrome_driver_path)

# Initialize the WebDriver
driver = webdriver.Chrome(service=service, options=chrome_options)

# Initialize Excel workbook and sheet
# wb = Workbook()
# wb.active.title = "Scraped Data"

# Open a webpage
driver.get("https://poit.bolagsverket.se/poit-app/")

wb = Workbook()
wb.active.title = "Scraped Data"

try:
    # cookie_button = WebDriverWait(driver, 10).until(
    #     EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Acceptera förvalda cookies']"))  # Adjust selector based on actual ID
    # )
    # cookie_button.click()

    element = WebDriverWait(driver, 40).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'a[href="/poit-app/sok"]'))
    )
    element.click()
    cookie_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Acceptera förvalda cookies']"))  # Adjust selector based on actual ID
    )
    cookie_button.click()
    ###########This is the part for getting URLs to be web scrapped.######################
    search_announce_btn = WebDriverWait(driver, 40).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[href="/poit-app/sok"]'))  # Adjust selector based on actual ID
    )
    search_announce_btn.click()

    period_selection_btn = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "tidsperiod"))
    )
    print(f"tidsperiod=period_selection_btn")
    period_selection_btn.set_keys('Annan period')

    # period_option_btn = WebDriverWait(driver, 10).until(
    #     EC.element_to_be_clickable((By.XPATH, "//select[@id='container_suggestions_caretype']//option[@value='ANNAN_PERIOD']"))
    # )
    # period_option_btn.click()


    # date_from_btn = WebDriverWait(driver, 10).until(
    #     EC.element_to_be_clickable((By.ID, "publiceringsdatum-from"))
    # )
    # date_from_btn.send_keys("2023-10-03")
    
    # date_from_btn = WebDriverWait(driver, 10).until(
    #     EC.element_to_be_clickable((By.ID, "publiceringsdatum-tom"))
    # )
    # date_to_btn.send_keys("2023-10-03")

    # show_more_btn = WebDriverWait(driver, 10).until(
    #     EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Visa fler sökalternativ')]"))
    # )
    # show_more_btn.click()

    # list_items = WebDriverWait(driver, 10).until(
    #     EC.presence_of_all_elements_located((By.CSS_SELECTOR, "li.categories-topic-list__item input[name='caretype-list']"))
    # )
    # list_values = [item.get_attribute('value') for item in list_items]
    
    # close_category_list_button = WebDriverWait(driver, 10).until(
    #     EC.element_to_be_clickable((By.CSS_SELECTOR, "button.category-lists-modal__modal-close"))
    # )
    # close_category_list_button.click()

    # total_list = []
    # list_values = [
    #     "Akutsjukvård",
    #     "Akutsjukvård, barn- och ungdomskirurgi",
    #     "Akutsjukvård, barn- och ungdomsmedicin",
    #     "Akutsjukvård, barn- och ungdomsortopedi",
    #     "Akutsjukvård, barn- och ungdomspsykiatri",
    #     "Akutsjukvård, gynekologi",
    #     "Akutsjukvård, infektion",
    #     "Akutsjukvård, internmedicin",
    #     "Akutsjukvård, kardiologi",
    #     "Akutsjukvård, kirurgi",
    #     "Akutsjukvård, käkkirurgi",
    #     "Akutsjukvård, neurologi",
    #     "Akutsjukvård, ortopedi",
    #     "Akutsjukvård, sexuellt våldsutsatta",
    #     "Akutsjukvård, vuxenpsykiatri",
    #     "Akutsjukvård, ögonsjukvård",
    #     "Akutsjukvård, öron-, näs- och halssjukvård"
    # ]

    # for value in list_values:
    #     category_input_field = WebDriverWait(driver, 10).until(
    #         EC.presence_of_element_located((By.ID, "caretype"))
    #     )

    #     ok_button = WebDriverWait(driver, 10).until(
    #         EC.element_to_be_clickable((By.CSS_SELECTOR, "button.findcare-form__button"))
    #     )

    #     print("-------------------------------------------------", value)
    #     try:
    #         clear_button = WebDriverWait(driver, 1).until(
    #             EC.element_to_be_clickable((By.CLASS_NAME, "findcare-searchfield__clear-field"))
    #         )
    #         if clear_button:
    #             clear_button.click()
    #     except (NoSuchElementException, TimeoutException):
    #         pass
        
    #     category_input_field.clear()

    #     category_input_field.send_keys(value)
    #     time.sleep(1)

    #     ok_button.click()
    #     time.sleep(1)
    #     ok_button.click()
        
        
    #     while True:
    #         try:
    #             # Wait until the parent div with class 'c-searchlist__footer' is present
    #             footer_div = WebDriverWait(driver, 10).until(
    #                 EC.presence_of_element_located((By.CSS_SELECTOR, "div.c-searchlist__footer"))
    #             )

    #             # Find the button within the 'c-searchlist__footer' div
    #             show_more_button = footer_div.find_element(By.CSS_SELECTOR, "button.c-pagination")

    #             # Click the "Visa 10 till" button
    #             driver.execute_script("arguments[0].scrollIntoView();", show_more_button)
    #             show_more_button.click()
    #             time.sleep(random.randint(4, 5))
    #         except (NoSuchElementException, TimeoutException):
    #             break
    #         except StaleElementReferenceException:
    #             continue

    #     try:
    #         info_divs = WebDriverWait(driver, 10).until(
    #             EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.findcare-search-list-item__info a"))
    #         )

    #         for item in info_divs:
    #             key = value
    #             link = item.get_attribute('href')
    #             total_list.append({'key': key, 'link': link})
    #     except Exception as e:
    #         print(f"Error==========: {e}")
    #     print(total_list)
    #     print(len(total_list))

    # # Print the links or do further processing
    # print("total=====================", total_list)
    # print("total count=====================",len(total_list))

    # # Save total list as a JSON file
    # with open('total_list.json', 'w') as json_file:
    #     json.dump(total_list, json_file, indent=4)

    # with open('total_list.json', 'r') as json_file:
    #     datas = json.load(json_file)

    # for data in datas:
    #     link = data.get('link')
    #     category = data.get('key')

    #     if link:
    #         print(f"Opening link: {link}")
    #         driver.get(link)

    #         random_time = random.randint(10, 15) / 10
    #         time.sleep(random_time)
    #         ws = wb.active

    #         headers = ['Name', 'Visiting Address', 'Visiting Area', 'County', 'Municipality', 'Postal Address', 'Postal Zip', 'Postal Area', 'Phone Number', 'URL', 'Category']

    #         # Write headers if the first row is empty
    #         if ws['A1'].value is None:
    #             for col, header in enumerate(headers, start=1):
    #                 cell = ws.cell(row=1, column=col, value=header)
    #                 cell.alignment = Alignment(horizontal='center', vertical='center')

    #         name = ''
    #         try:
    #             heading_element = WebDriverWait(driver, 10).until(
    #                 EC.presence_of_element_located((By.CSS_SELECTOR, "h1.contact-header__heading"))
    #             )
    #             name = heading_element.text
    #         except (NoSuchElementException, TimeoutException):
    #             pass
            
    #         visiting_address = ''
    #         visiting_area = ''
    #         try:
    #             address_dt = WebDriverWait(driver, 10).until(
    #                 EC.presence_of_element_located((By.XPATH, "//dl[@class='find-us__address']//dt[text()='Besöksadress:']"))
    #             )
    #             address_dd = address_dt.find_element(By.XPATH, "following-sibling::dd")
    #             address_text = address_dd.text
    #             address_parts = address_text.split(",")
    #             address_parts = [part.strip() for part in address_parts]
    #             if len(address_parts) > 0:
    #                 visiting_address = address_parts[0]
    #             if len(address_parts) > 1:
    #                 visiting_area = address_parts[1]
    #         except (NoSuchElementException, TimeoutException):
    #             pass

    #         county = ''
    #         municipality = ''
    #         try:
    #             lan_kommun_dt = WebDriverWait(driver, 10).until(
    #                 EC.presence_of_element_located((By.XPATH, "//dl[@class='find-us__address']//dt[text()='Län, kommun:']"))
    #             )
    #             lan_kommun_dd = lan_kommun_dt.find_element(By.XPATH, "following-sibling::dd")
    #             lan_kommun_text = lan_kommun_dd.text
    #             lan_kommun_parts = lan_kommun_text.split(",")
    #             lan_kommun_parts = [part.strip() for part in lan_kommun_parts]
    #             if len(lan_kommun_parts) > 0:
    #                 county = lan_kommun_parts[0]
    #             if len(lan_kommun_parts) > 1:
    #                 municipality = lan_kommun_parts[1]
    #         except (NoSuchElementException, TimeoutException):
    #             pass
                
    #         postal_address = ''
    #         postal_zip = ''
    #         postal_area = ''
    #         try:
    #             post_address_dt = WebDriverWait(driver, 10).until(
    #                 EC.presence_of_element_located((By.XPATH, "//dt[text()='Postadress:']/following-sibling::dd[1]"))
    #             )

    #             # Extract the text of the post address
    #             post_address_text = post_address_dt.text
    #             pattern = r"^(.*?),\s*(\d{3}\s*\d{2})\s*(.*)$"
    #             match = re.match(pattern, post_address_text)
    #             if match:
    #                 postal_address = match.group(1)  # Group 1 is the full postal address
    #                 postal_zip = match.group(2)      # Group 2 is the postal zip code
    #                 postal_area = match.group(3)     # Group 3 is the postal area
    #             else:
    #                 print("Pattern did not match the address format")
    #         except (NoSuchElementException, TimeoutException):
    #             pass

    #         try:
    #             phone_number_element = WebDriverWait(driver, 10).until(
    #                 EC.visibility_of_all_elements_located((By.CSS_SELECTOR, '.phonehours__numbers__list__item__link'))
    #             )
    #             if phone_number_element:
    #                 phone_number = phone_number_element[0].text
    #             else:
    #                 phone_number = ''
    #         except (NoSuchElementException, TimeoutException):
    #             phone_number = ''

    #         web_url = ''
    #         try:
    #             webpage_element = WebDriverWait(driver, 10).until(
    #                 EC.presence_of_element_located((By.CSS_SELECTOR, '.find-us__webpage'))
    #             )
    #             web_url = webpage_element.get_attribute('href')
    #         except:
    #             print("Web URL not found.")
            
    #     else:
    #         print("Link is missing in the JSON data")

    #     row = ws.max_row + 1

    #     # Write other product information to adjacent cells
    #     ws[f'A{row}'] = name  # Assigning the extracted value or default value
    #     ws[f'B{row}'] = visiting_address  # Assigning the extracted value or default value
    #     ws[f'C{row}'] = visiting_area
    #     ws[f'D{row}'] = county
    #     ws[f'E{row}'] = municipality
    #     ws[f'F{row}'] = postal_address
    #     ws[f'G{row}'] = postal_zip
    #     ws[f'H{row}'] = postal_area
    #     ws[f'I{row}'] = phone_number
    #     ws[f'J{row}'] = web_url
    #     ws[f'K{row}'] = category

    #     print(f"Product information for {name} added to the Excel file.")

    #     # Set the alignment of all cells to center both horizontally and vertically
    #     for row in ws.iter_rows():
    #         for cell in row:
    #             cell.alignment = Alignment(horizontal='center', vertical='center')

    # # # Save Excel file
    # wb.save('product_info.xlsx')
    # print('Your excel file exported successfully!')

except Exception as e:
    print(f"Error-=-=-=-=: {e}")
time.sleep(1000)  # Ensure the page loads completely